from openpyxl import Workbook
from paquetes import myframe
import datetime, csv, time
import sys

#archivo de plataforma no guarda los datos de las zondas
# que no se reportan! hay que rellenar!

#Nota: para acceder a las celdas, primero toca crearlas, 
#cuando una hoja de trabajo es creada no contiene celdas.
# Formatear la fecha y la hora para el nombre del archivo
archivo = sys.argv[1]
dias = ['lunes','martes','miercoles','jueves','viernes','sabado']
libro = Workbook() #instanciar un libro de excel
hoja = libro.get_active_sheet() #activar una hoja de calculo
hoja.title ="Bajada"

id_muestra = hoja.cell("A1")
id_muestra.value="id_muestra"
rbd = hoja.cell("B1")
rbd.value="rbd"
id_enlace = hoja.cell("C1")
id_enlace.value="id_enlace"
anexo = hoja.cell("D1")
anexo.value="anexo"
establecimiento = hoja.cell("E1")
establecimiento.value="establecimiento"
tip_servicio = hoja.cell("F1")
tip_servicio.value="tip_servicio"
velocidad = hoja.cell("G1")
velocidad.value="velocidad"
id_bts = hoja.cell("H1")
id_bts.value="id_bts"
nombre_bts = hoja.cell("I1")
nombre_bts.value="nombre_bts"
fecha_muestra = hoja.cell("J1")
fecha_muestra.value="fecha_muestra"
dia_muestra = hoja.cell("K1")
dia_muestra.value="dia_muestra"
hora_muestra = hoja.cell("L1")
hora_muestra.value="hora_muestra"
valor_muestra = hoja.cell("M1")
valor_muestra.value="valor_muestra"
detalle_vel_bajada = 'DETALLE_VEL_BAJADA_CTR'
detalle_vel_subida = 'DETALLE_VEL_SUBIDA_CTR'
detalle_Disponibilidad = 'DETALLE_DISPO_CTR'
fecha_actual = datetime.date.today().strftime("%Y%m%d")
detalle_vel_bajada=fecha_actual +'_'+ detalle_vel_bajada +'.1'
detalle_vel_subida=fecha_actual +'_'+ detalle_vel_subida +'.1'
detalle_Disponibilidad=fecha_actual +'_'+ detalle_Disponibilidad +'.1'
idm = 1
dd = 0
horas = 8
myframe.cls()
# Cargar archivo CSV
print archivo
reader = csv.reader(open(archivo,'rb'),delimiter=",")
for index,row in enumerate(reader):
	id_muestra = hoja.cell("A"+str(index+2))
	id_muestra.value=idm

	rbd=hoja.cell("B"+str(index+2))
	rbd.value=row[3]

	id_enlace=hoja.cell("C"+str(index+2))
	id_enlace.value=row[1]

	anexo = hoja.cell("D"+str(index+2))
	anexo.value=row[2]

	establecimiento=hoja.cell("E"+str(index+2))
	establecimiento.value=row[0]

	tip_servicio=hoja.cell("F"+str(index+2))
	tip_servicio.value="RADIOFRECUENCIA"

	velocidad=hoja.cell("G"+str(index+2))
	velocidad.value="1024"

	id_bts = hoja.cell("H"+str(index+2))
	id_bts.value="NA"

	nombre_bts=hoja.cell("I"+str(index+2))
	nombre_bts.value="NA"

	fecha_muestra=hoja.cell("J"+str(index+2))
	fec = str(row[10])
	fec=fec.split()
	fecha_muestra.value=fec[0]

	dia_muestra = hoja.cell("K"+str(index+2))
	if dd == 6:
		dd = 0
	dia_muestra.value=dias[dd]

	hora_muestra=hoja.cell("L"+str(index+2))
	hora_muestra.value=horas
	
	hor=fec[1].split(":")
	print hor[0] + " - " + str(horas)

	if int(hor[0]) == horas:
		print fec[1]
		valor_muestra=hoja.cell("M"+str(index+2))
		valor_muestra.value=row[12]
	# is trim(row[11])=horas
	# escribir la hora si no NA
	idm += 1
	horas +=1
	if idm == 12:
		idm = 1
	if horas == 19:
		dd += 1
		horas = 8
	#print 'personas'+str(index+1)
 	#print 'Nombre: ' + row[0] + ', Apellidos: ' + row[1] + ', Edad: ' + row[2] + '\r'
#hoja2 = libro.create_sheet(0) #Insertar una segunda hoja
#hoja = libro.create_sheet(title="Nueva hoja")
#Guardar la hoja de calculo
libro.save(detalle_vel_bajada+'.xlsx') 