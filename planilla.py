from openpyxl import Workbook
import datetime, csv

#Nota: para acceder a las celdas, primero toca crearlas, 
#cuando una hoja de trabajo es creada no contiene celdas.
# Formatear la fecha y la hora para el nombre del archivo

year=datetime.date.today().strftime("%Y")
mes=datetime.date.today().strftime("%B")
libro = Workbook() #instanciar un libro de excel
hoja = libro.get_active_sheet() #activar una hoja de calculo
hoja.title ="Disponibilidad"

id_muestra = hoja.cell("A1")
id_muestra.value="id_muestra"
rbd = hoja.cell("B1")
rbd.value="rbd"
id_enlace = hoja.cell("C1")
id_enlace.value="id_enlace"
anexo = hoja.cell("D1")
anexo.value="anexo"
establecimiento = hoja.cell("E1")
establecimiento.value="establecimiento"
tip_servicio = hoja.cell("F1")
tip_servicio.value="tip_servicio"
id_bts = hoja.cell("G1")
id_bts.value="id_bts"
nombre_bts = hoja.cell("H1")
nombre_bts.value="nombre_bts"
fecha_muestra = hoja.cell("I1")
fecha_muestra.value="fecha_muestra"
dia_muestra = hoja.cell("J1")
dia_muestra.value="dia_muestra"
hora_muestra = hoja.cell("K1")
hora_muestra.value="hora_muestra"
disponibilidad = hoja.cell("L1")
disponibilidad.value="disponibilidad"

#celda.value = "hola hola"
# Cargar archivo CSV
reader = csv.reader(open('prueba.csv','rb'),delimiter=",")
for index,row in enumerate(reader):

	Host = hoja.cell("A"+str(index+2))
	Host.value=row[0]

	IdServicio=hoja.cell("B"+str(index+2))
	Apellidos.value=row[1]

	Apellidos=hoja.cell("B"+str(index+2))
	Apellidos.value=row[1]
	#print 'personas'+str(index+1)
 	#print 'Nombre: ' + row[0] + ', Apellidos: ' + row[1] + ', Edad: ' + row[2] + '\r'

#hoja2 = libro.create_sheet(0) #Insertar una segunda hoja
#hoja = libro.create_sheet(title="Nueva hoja")
#Guardar la hoja de calculo
libro.save(mes+'_'+year+'.xlsx') 